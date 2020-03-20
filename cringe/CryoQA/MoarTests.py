#cd ~/gitrepo/nist_lab_internals/viper/cringe/tower/
import sys
import time
from PyQt4 import QtGui, QtCore
from PyQt4.QtCore import Qt, SIGNAL
from PyQt4.QtGui import QWidget, QDoubleSpinBox, QSpinBox, QFrame, QGroupBox,QToolButton, QPushButton, QSlider, QMenu

# import named_serial
import struct
import towerwidget

import numpy as np
import easyClient
import matplotlib.pyplot as plt
import openpyxl as xl

app = QtGui.QApplication(sys.argv)
app.setStyle("plastique")
app.setStyleSheet("""    QPushbutton{font: 10px; padding: 6px}
                            QToolButton{font: 10px; padding: 6px}""")

widget = towerwidget.TowerWidget(nameaddrlist=['SAB1', '1', 'SAFB', '15', 'SQ1FB', '11'])
import numpy as np
import easyClient
import matplotlib.pyplot as plt

conversion = 1.0/16384.0*1.0/2.4/2.10*1000.0

c = easyClient.EasyClient()
c.setupAndChooseChannels()

class CryoQA(object):

    def __init__(self, rows, columns, channel, chipID, filename, SQ1bStart=1, SQ1bEnd=20000, SQ1bSteps=100, startIndx=0
                 , pointsPerSlice=2048):


        self.rows = rows
        self.columns = columns
        self.start = SQ1bStart
        self.end = SQ1bEnd
        self.steps = SQ1bSteps
        self.startIndx = startIndx
        self.pointsPerSlice = pointsPerSlice
        self.channel = channel
        self.chipID = chipID
        self.filename = filename

    def SQ1bStep(self):
    #
    #Creates an array of values to step tower bias through.
    #
        return np.linspace(self.start, self.end, self.steps)

    def listMaker(self):
    #
    #Creates an empty list to append min, max, max-min values to.
    #
        a = [[[] for i in range(0, self.rows)] for j in range(0, self.columns)]
        return a

    def biasStepper(self, val):
    #
    #Sets tower bias to a given value (val).
    #
        if self.channel > 7:
            widget.towercards['SQ1FB'].towerchannels[self.channel-8].dacspin.setValue(val)
        else:
            widget.towercards['SQ1FB'].towerchannels[self.channel].dacspin.setValue(val)

    def averager(self):
    #
    # Gathers data from easyClient and averages one period of it over as many periods as possible in the data.
    # Returns two arrays: fb, err
    #
        data = c.getNewData(minimumNumPoints=20480, exactNumPoints=True)

    def SQ1bStep(self):
        return np.arange(self.start, self.end, self.steps)

    def listMaker(self):
        a = [[[] for i in range(0, self.rows)] for j in range(0, self.columns)]
        return a

    #def biasStepper(self, val):
        #do some stuff I don't quite know how to do yet
        #follows Galen's code:
        #towercard = self.mm.cringe.tower_widget.towercards["SQ1b"]
        #for towerchannel in towercard.towerchannels:
            #towerchannel.dacspin.setValue(val)


    def test(self):
        a = self.listMaker()
        for element in self.SQ1bStep():
            #self.biasStepper(element)
            fb, err = self.averager()
            print "Toot!"
            for i, row in enumerate(fb):
                min = np.amin(row)
                max = np.amax(row)
                a[0][i].append([min, max])
        return a


    def averager(self):
        data = c.getNewData()

        slices = np.int_(data.shape[2] / self.pointsPerSlice)
        start = self.startIndx
        end = start + self.pointsPerSlice

        i = 0
        fb = np.zeros(self.pointsPerSlice)

        while i != slices:
            fb = fb + data[0, :, start:end, 1]
            start += self.pointsPerSlice
            end += self.pointsPerSlice
            i += 1

        return fb / slices, data[0, :, self.startIndx:self.startIndx + self.pointsPerSlice, 0]

    def baseline(self):
    #
    # Measures the baseline squid curve value before any bias is applied.
    #
        self.biasStepper(1)
        fb, err = self.averager()
        minMax = np.amax(fb)-np.amin(fb)
        return fb*conversion

    def maxMin(self):
    #
    #Steps tower bias and runs averager over every step.
    #Returns [col, row, data points, min=0/max=1/max-min=3]
    #
        a = self.listMaker()
        for element in self.SQ1bStep():
            self.biasStepper(element)
            fb, err = self.averager()
            for i, row in enumerate(fb):
                min = np.amin(row)
                max = np.amax(row)
                a[0][i].append([(min*conversion), (max*conversion),
                                ((max-min)*conversion)])
        return np.array(a)

    def findIcs(self):
    #
    #Pulls Ic values out of data.
    #
        Baseline = self.baseline()
        array = self.maxMin()
        icMaxArray = np.zeros(len(array[0, :, 0, 2]))
        icMinArray = np.zeros(len(array[0, :, 0, 2]))

        for i, maxElement in enumerate(array[0, :, :, 2]):
            icMax = np.amax(maxElement)
            icMaxArray[i] = icMax

        for j, minElement in enumerate(array[0, :, :, 2]):
            base = np.amax(Baseline[j]) - np.amin(Baseline[j])
            #print base
            #print minElement
            sigma = np.std(Baseline[j])
            #print sigma
            icMinIndex = np.where((minElement - base) >= 14.0*sigma)
            #print minElement - base
            #print icMinIndex
            icMinIndex = icMinIndex[0][0]
            icMinArray[j] = minElement[icMinIndex]

        return icMinArray, icMaxArray

    def saveToExcel(self):
    #
    #Puts icMin/icMax data into an excel file. NOTE: the data is only saved if all 8 channels are measured, so if you
    #need save the file before that, you must manually save it by typing wb.save("filename".xls).
    #Also, for now, you have to call the function for each channel by creating a new instance of the class.
    #
        icMin, icMax = self.findIcs()

        if self.channel == 0:
            wb = xl.Workbook()
            activeBook = wb.active

            activeBook["A1"] = "Chip ID"
            activeBook["B1"] = "Row"
            activeBook["C1"] = "Icmin (uA)"
            activeBook["D1"] = "Icmax (uA)"

            activeBook["A" + str((self.channel*self.rows) + 2)] = self.chipID

            for element in np.arange(0, self.rows):
                activeBook["B" + str((self.channel*self.rows)+element+2)] = str(element)

            for i, element in enumerate(icMin):
                activeCell = activeBook.cell(row=(self.channel*self.rows)+i+2, column=3)
                activeCell.value = icMin[i]
                activeCell = activeBook.cell(row=(self.channel*self.rows)+i+2, column=4)
                activeCell.value = icMax[i]

            activeBook.title = raw_input("Input title name:")
            wb.save("/home/pcuser/Documents/Squid_Screening_Data/MUX15/" + self.filename)

        else:
            wb = xl.load_workbook("/home/pcuser/Documents/Squid_Screening_Data/MUX15/" + self.filename)
            activeBook = wb.active

            activeBook["A" + str((self.channel * self.rows) + 2)] = self.chipID

            for element in np.arange(0, self.rows):
                activeBook["B" + str((self.channel * self.rows) + element + 2)] = str(element)

            for i, element in enumerate(icMin):
                activeCell = activeBook.cell(row=(self.channel * self.rows) + i + 2, column=3)
                activeCell.value = icMin[i]
                activeCell = activeBook.cell(row=(self.channel * self.rows) + i + 2, column=4)
                activeCell.value = icMax[i]

            wb.save("/home/pcuser/Documents/Squid_Screening_Data/MUX15/" + self.filename)



#a = CryoQA(12, 1, 4, "Row 4 Col 0", "Test_7.xlsx")
#b = a.listMaker()
#print b[0][0]
#icMin, icMax = a.findIcs()
#print cool
#plt.plot(np.linspace(1, 40000, 160), cool[0, 0, :, 0], '.')
#plt.plot(np.linspace(1, 40000, 160), cool[0, 0, :, 1], '.')x
#plt.show()
#print icMin, icMax
#a.saveToExcel()
#=======

a = CryoQA(12, 1, 15, "4-10", "MUX15B_2017_01_13_Mrs_Multipurpose_10.xlsx")
a.saveToExcel()

#>>>>>>> 715295b10161e5229efd597f25d5092e576767d6
