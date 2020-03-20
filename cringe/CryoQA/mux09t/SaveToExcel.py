import sys
import argparse
import IPython
import glob

import numpy as np
import pylab as pl
import openpyxl as xl
import time

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Take data from ROACH2, demodulate it, and write the demodulated data to a file. Can also display various plots.')
    parser.add_argument('list_of_files',
                        type=str,
                        help='A list of paths to some files that will be parsed and plotted')
    parser.add_argument('-i',
                        dest='interactive',
                        action='store_true',
                        help='Enter interactive mode at end of script')
    args = parser.parse_args()

    if args.interactive:
        pl.ion()

    ListOfFiles = args.list_of_files
    fnames = glob.glob(str(args.list_of_files))
    fnames.sort()
    ic_mins = np.zeros((len(fnames), 33), 'int')
    ic_maxs = np.zeros((len(fnames), 33), 'int')
    data_sets = [np.load(f) for f in fnames]

    # tower scaling parameters
    tDACfs = 2**16
    tDACref = 2.5
    tRbias = 5100
    tscale = 1e6
    tfactor = tDACref*tscale/(tDACfs*tRbias)

    # BAD16 scaling parameters
    bDACfs = 2 ** 14
    bDACref = 1.0
    bRbias = 2000
    bscale = 1e6
    bfactor = bDACref * bscale / (bDACfs * bRbias)

    wb = xl.Workbook()
    ws = wb.active
    ws.title = "info"

    ws['B1'] = "BIAS"

    ws['A2'] = "DACfs"
    ws['A3'] = "DACref"
    ws['A4'] = "Rbias"
    ws['A5'] = "scale"
    ws['A6'] = "multiplier"

    ws['B2'] = bDACfs
    ws['B3'] = bDACref
    ws['B4'] = bRbias
    ws['B5'] = bscale
    ws['B6'] = bfactor

    ws['C2'] = "bits"
    ws['C3'] = "V"
    ws['C4'] = "ohms"
    ws['C5'] = "uA/A"
    ws['C6'] = "uA/bit"

    ws['A8'] = "source"
    ws['B8'] = ListOfFiles

    ws1 = wb.create_sheet(title="Icmin")
    ws2 = wb.create_sheet(title="Icmax")

    today = time.localtime()
    now = str(today.tm_year) + '_' + str(today.tm_mon) + '_' + str(today.tm_mday) + '_' + str(today.tm_hour) + str(today.tm_min)
    basefilepath = "/home/pcuser/Documents/cdr_data/mux09t/"

    for i, file_path in enumerate(fnames):

        ws.cell(row=10+i, column=1, value=fnames[i])
        last_part_fname = fnames[i].rsplit('/', 1)[-1]
        ic_mins = data_sets[i]['icmin']
        ic_maxs = data_sets[i]['icmax']
        ws1.cell(row=1, column=i+2, value=last_part_fname.rsplit('_')[0])
        ws2.cell(row=1, column=i+2, value=last_part_fname.rsplit('_')[0])
        if i==0:
            ws1.cell(row=1, column=1, value='row')
            ws2.cell(row=1, column=1, value='row')

        for j in range(len(ic_mins)):
            if i==0:
                if j==0:
                    ws1.cell(row=j+2, column=1, value='DS')
                    ws2.cell(row=j+2, column=1, value='DS')
                else:
                    ws1.cell(row=j+2, column=i+1, value=j-1)
                    ws2.cell(row=j+2, column=i+1, value=j-1)

            ws1.cell(row=j+2, column=i+2, value=float(ic_mins[j])*bfactor)
            ws1.cell(row=j+2, column=i+2).number_format = '####.#'
            ws2.cell(row=j+2, column=i+2, value=float(ic_maxs[j])*bfactor)
            ws2.cell(row=j+2, column=i+2).number_format = '####.#'

    filename = raw_input("Please enter file identifier:")
    path = basefilepath + filename + '_' + now + '.xls'

    wb.save(path)

            # '''Make a plot to show the Ic Sweeps for each row overlayed'''
            # for j in range(data_sets[i]['row_average_adcpp_sweep'].shape[0]):
            #     pl.title('Row Bias Sweep :' + last_part_fname)
            #     pl.plot(data_sets[i]['row_sweep_dac_values'], data_sets[i]['row_average_adcpp_sweep'][j, :],
            #             label=('Row:' + str(j)),
            #             color=colors_row[j])
            #     pl.xlabel('BAD16 DAC Value')
            #     pl.ylabel('SQ1 Modulation Depth [arbs]')
            #     pl.legend()
            # '''Make a plot showing IC_min and IC_max overlayed for each file'''
            # pl.figure()
            # pl.title('Icmin/max :' + last_part_fname)
            # pl.plot(data_sets[i]['icmin'], 'r', label='Ic_min')
            # pl.plot(data_sets[i]['icmax'], 'b', label='Ic_max')
            # pl.xlabel('Row Number')
            # pl.ylabel('BAD16 Bias value')
            # pl.legend()

    # else:
    #     pl.figure()
    #     for i in range(len(fnames)):
    #         ic_maxs[i] = data_sets[i]['icmax']
    #         pl.plot(ic_maxs[i], '-*', label=str(data_sets[i]['chip_id']), color=colors_col[i])
    #
    #     pl.legend()
    #     pl.xlabel('Row Number')
    #     pl.ylabel('DAC Bias Current')

    if args.interactive:
        IPython.embed()
    # else:
        # pl.show()
        # return
