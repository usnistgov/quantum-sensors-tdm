#cd ~/gitrepo/nist_lab_internals/viper/cringe/tower/
import sys
import time
from PyQt4 import QtGui, QtCore
from PyQt4.QtCore import Qt, SIGNAL
from PyQt4.QtGui import QWidget, QDoubleSpinBox, QSpinBox, QFrame, QGroupBox,QToolButton, QPushButton, QSlider, QMenu

# import named_serial
import struct
import towerwidget

import numpy as np
import easyClient
import matplotlib.pyplot as plt
import openpyxl as xl

app = QtGui.QApplication(sys.argv)
app.setStyle("plastique")
app.setStyleSheet("""    QPushbutton{font: 10px; padding: 6px}
                            QToolButton{font: 10px; padding: 6px}""")

#widget = towerwidget.TowerWidget(nameaddrlist=['SAB1', '1', 'SAFB', '15', 'SQ1FB', '11'])


class CryoQA(object):
    def __init__(self, rows, columns, channel, chipID, filename, SQ1bStart=1, SQ1bEnd=20000, SQ1bSteps=200, startIndx=0,
                 pointsPerSlice=2048):

        self.c = easyClient.EasyClient()
        self.c.setupAndChooseChannels()
        self.widget = towerwidget.TowerWidget(nameaddrlist=['SAB1', '1', 'SAFB', '15', 'SQ1FB', '11'])

        self.rows = rows
        self.columns = columns
        self.start = SQ1bStart
        self.end = SQ1bEnd
        self.steps = SQ1bSteps
        self.startIndx = startIndx
        self.biasvals_I = None
        self.pointsPerSlice = pointsPerSlice
        self.channel = channel
        self.chipID = chipID
        self.filename = filename

        #   MrsMP PCB parameters
        self.Rsh = 1.0                      # [ohms]    cold shunt,      guess => replace with accurate measure

        #   Series array parameters
        self.Msafb = 43.8e-12               # [H]       from Mrs_Multipurpose_11_Row_MUX_Testing_Module.xls
        self.Msain = 103.5e-12              # [H]       from Mrs_Multipurpose_11_Row_MUX_Testing_Module.xls
        self.Msar = self.Msain/self.Msafb

        #   MUX15b parameters
        self.Msq1fb = 10e-12                # [H]       guess
        self.Msq1in = 676e-12               # [H]       measured 10/11/16 for MUX15b

        #   crate parameters
        self.crateDACdr = 2.0**14           # [bits]
        self.crateDACvr = 1.0               # [V]

        #   tower card parameters
        self.towerDACdr = 2.0**16           # [bits]
        self.towerDACvr = 2.5               # [V]
        self.Rsq1bias = 5100.0              # [ohms]    BIAS card
        self.Rsafb = 5100.0                 # [ohms]    BIAS card
        self.Rsq1fb = 2000.0                # [ohms]    SMB FTHRU card
        self.Rin = 20000                    # [ohms]    modified HDMI FTHRUs

        #   current scaling parameters
        self.fbscale = 1/self.crateDACdr*self.crateDACvr/self.Msar/self.Rsafb
        self.BiasScale = 1/self.towerDACdr*self.towerDACvr/self.Rsq1bias



    def SQ1bStep(self):
    #
    #Creates an array of values to step tower bias through.
    #
        return np.linspace(self.start, self.end, self.steps)

    def listMaker(self):
    #
    #Creates an empty list to append min, max, max-min values to.
    #
        a = [[[] for i in range(0, self.rows)] for j in range(0, self.columns)]
        return a

    def biasStepper(self, val):
    #
    #Sets tower bias to a given value (val).
    #
        self.widget.towercards['SQ1FB'].towerchannels[self.channel].dacspin.setValue(val)

    def averager(self):
    #
    # Gathers data from easyClient and averages one period of it over as many periods as possible in the data.
    # Returns two arrays: fb, err
    #
        data = self.c.getNewData(minimumNumPoints=self.pointsPerSlice*50, exactNumPoints=True)

        slices = np.int_(data.shape[2] / self.pointsPerSlice)
        start = self.startIndx
        end = start + self.pointsPerSlice

        i = 0
#        self.biasStepper(bias)
        fb = np.zeros(self.pointsPerSlice)
        err = np.zeros(self.pointsPerSlice)

        while i != slices:
            fb = fb + data[0, :, start:end, 1]
            err = err + data[0, :, start:end, 0]
            start += self.pointsPerSlice
            end += self.pointsPerSlice
            i += 1

        return fb * self.fbscale / slices, err / slices

    def baseline(self, bias):
    #
    # Measures the baseline squid curve value before any bias is applied.
    #
        self.biasStepper(bias)
        fb, err = self.averager()
        return fb

    def maxMin(self):
    #
    #Steps tower bias and runs averager over every step.
    #Returns [col, row, data points, min=0/max=1/max-min=3]
    #
        a = self.listMaker()

        biasvals = self.SQ1bStep()
        self.biasvals_I = biasvals * self.BiasScale
#        for element in self.SQ1bStep():
        for sample in biasvals:
            self.biasStepper(biasvals[sample])
            fb, err = self.averager()
            for i, row in enumerate(fb):
                min = np.amin(row)
                max = np.amax(row)
                a[0][i].append([min, max, (max-min)])

        return np.array(a)

    def findIcs(self):
    #
    #Pulls Ic values out of data.
    #
        Baseline = self.baseline()
        array = self.maxMin()
        icMaxArray = np.zeros(len(array[0, :, 0, 2]))
        icMinArray = np.zeros(len(array[0, :, 0, 2]))

        for i, maxElement in enumerate(array[0, :, :, 2]):
            icMax = np.amax(maxElement)
            icMaxArray[i] = icMax
            print icMax

        for j, minElement in enumerate(array[0, :, :, 2]):
            base = np.amax(Baseline[j]) - np.amin(Baseline[j])
            #print base
            #print minElement
            sigma = np.std(Baseline[j])
            #print sigma
            icMinIndex = np.where((minElement - base) >= 14*sigma)
            #print minElement - base
            print icMinIndex
            icMinIndex = icMinIndex[0][0]
            icMinArray[j] = minElement[icMinIndex]

        return icMinArray, icMaxArray

    def saveToExcel(self):
    #
    #Puts icMin/icMax data into an excel file. NOTE: the data is only saved if all 8 channels are measured, so if you
    #need save the file before that, you must manually save it by typing wb.save("filename".xls).
    #Also, for now, you have to call the function for each channel by creating a new instance of the class.
    #
        icMin, icMax = self.findIcs()

        if self.channel == 0:
            wb = xl.Workbook()
            activeBook = wb.active

            activeBook["A1"] = "Chip ID"
            activeBook["B1"] = "Row"
            activeBook["C1"] = "Icmin (uA)"
            activeBook["D1"] = "Icmax (uA)"

            activeBook["A" + str((self.channel*self.rows) + 2)] = self.chipID

            for element in np.arange(0, self.rows):
                activeBook["B" + str((self.channel*self.rows)+element+2)] = str(element)

            for i, element in enumerate(icMin):
                activeCell = activeBook.cell(row=(self.channel*self.rows)+i+2, column=3)
                activeCell.value = icMin[i]
                activeCell = activeBook.cell(row=(self.channel*self.rows)+i+2, column=4)
                activeCell.value = icMax[i]

            activeBook.title = raw_input("Input title name:")
            wb.save("/home/pcuser/Documents/script_testing/" + self.filename)

        else:
            wb = xl.load_workbook("/home/pcuser/Documents/script_testing/" + self.filename)
            activeBook = wb.active

            activeBook["A" + str((self.channel * self.rows) + 2)] = self.chipID

            for element in np.arange(0, self.rows):
                activeBook["B" + str((self.channel * self.rows) + element + 2)] = str(element)

            for i, element in enumerate(icMin):
                activeCell = activeBook.cell(row=(self.channel * self.rows) + i + 2, column=3)
                activeCell.value = icMin[i]
                activeCell = activeBook.cell(row=(self.channel * self.rows) + i + 2, column=4)
                activeCell.value = icMax[i]

            wb.save("/home/pcuser/Documents/script_testing/" + self.filename)



#a = CryoQA(12, 1, 4, "Row 4 Col 0", "Test_7.xlsx")
#b = a.listMaker()
#print b[0][0]
#icMin, icMax = a.findIcs()
#print cool
#plt.plot(np.linspace(1, 40000, 160), cool[0, 0, :, 0], '.')
#plt.plot(np.linspace(1, 40000, 160), cool[0, 0, :, 1], '.')
#plt.show()
#print icMin, icMax
#a.saveToExcel()
